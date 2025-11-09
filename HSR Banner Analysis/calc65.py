import numpy as np
from calclib import *
from drawlib import *
import matplotlib.pyplot as plt
plt.rcParams["font.family"]="SimHei"

fig = plt.figure(num=1, figsize=(1,2))
pList = [0.1, 0.5, 0.9, 0.95, 0.99]

def calcQuantilesAndExpects() :
    global pList
    chaDist = calcEachNumDist(
        calcDistBigGrt(
            calcDist(
                genBaseProbArray(90, 0.006, 74, 0.06))) ,7)
    weaDist = calcEachNumDist(
        calcDistBigGrt(
            calcDist(
                genBaseProbArray(80, 0.008, 66, 0.07))) ,5)
    lp = len(pList)
    ans = np.zeros([8,6,lp], dtype=np.int32)
    expects = np.zeros([8,6], dtype=np.float64)
    for cn in range(8) :
        for wn in range(6) :
            if cn == 0 and wn == 0:
                continue
            elif cn == 0 :
                dist = weaDist[wn-1]
            elif wn == 0:
                dist = chaDist[cn-1]
            else:
                dist = calcConv(chaDist[cn-1], weaDist[wn-1])
            pi = 0
            sum = 0.0
            for i in range(dist.shape[0]) :
                sum += dist[i]
                if(pi<lp and pList[pi]<=sum) :
                    ans[cn][wn][pi] = i
                    pi += 1
                elif pi>=lp :
                    break
            expects[cn, wn] = calcExpect(dist)
    return ans, expects

from openpyxl import Workbook, worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def adjust_column_width(ws) :
    for ncol in range(1, ws.max_column+1) :
        max_length = 0
        column = get_column_letter(ncol)
        for nrow in range(1, ws.max_row+1) :
            try:
                cell = ws.cell(row=nrow, column=ncol)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        if max_length > 0 :
            ws.column_dimensions[column].width = min(max_length+2, 50) # 最大列宽50

def export_ans_to_excel(ans, expects, filename='output.xlsx'):
    """
    将三维数组ans导出为Excel表格
    
    参数:
    ans: 形状为[7,5,6]的三维numpy数组
    pList: 长度为6的列表，包含分位数数值
    filename: 输出文件名
    """
    global pList
    # 创建Excel工作簿
    wb = Workbook()
    # 删除默认创建的工作表
    wb.remove(wb.active)
    
    # 定义字体样式
    title_font = Font(name='黑体', bold=True, size=12)
    header_font = Font(name='宋体', bold=True, size=11)
    data_font = Font(name='Times New Roman', size=11)
    
    # 居中对齐
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 为每个第三维度创建表格
    lp = len(pList)
    for pi in range(lp):
        # 创建新的工作表
        sheet_name = f"Table_{pi+1}"
        ws = wb.create_sheet(title=sheet_name)
        
        # 设置表格标题
        title = f"分位数p={pList[pi]}"
        ws.merge_cells(f'A1:G1')  # 合并第一行的所有列
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # 设置列标题
        col_headers = [''] + [f'{ncol}精' for ncol in range(6)]
        for col_idx, header in enumerate(col_headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align
        
        # 设置行标题和数据
        for nrow in range(8):
            # 行标题
            row_header_cell = ws.cell(row=nrow+3, column=1)
            row_header_cell.value = f'{nrow-1}命'
            row_header_cell.font = header_font
            row_header_cell.alignment = center_align
            
            # 数据
            for ncol in range(6):
                data_cell = ws.cell(row=nrow+3, column=ncol+2)
                data_cell.value = int(ans[nrow, ncol, pi])  # 转换为int，因为dtype是int32
                data_cell.font = data_font
                data_cell.alignment = center_align
        
        # 通过枚举单元格，调整列宽
        adjust_column_width(ws)

        # 在表格之间添加空行（在下一个表格前添加）
        if pi < 5:  # 不是最后一个表格
            ws.append([])
    
    # 写期望表
    ws = wb.create_sheet(title="Expects")
    title = f"数学期望"
    ws.merge_cells(f'A1:G1')  # 合并第一行的所有列
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

     # 设置列标题
    col_headers = [''] + [f'{ncol}精' for ncol in range(6)]
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
    
    # 设置行标题和数据
    for nrow in range(8):
        # 行标题
        row_header_cell = ws.cell(row=nrow+3, column=1)
        row_header_cell.value = f'{nrow-1}命'
        row_header_cell.font = header_font
        row_header_cell.alignment = center_align
        
        # 数据
        for ncol in range(6):
            data_cell = ws.cell(row=nrow+3, column=ncol+2)
            data_cell.value = int(expects[nrow, ncol])  # 转换为int，因为dtype是int32
            data_cell.font = data_font
            data_cell.alignment = center_align
    adjust_column_width(ws)

    # 保存文件
    wb.save(filename)
    print(f"文件已保存为: {filename}")

if __name__ == "__main__" :
    ans, expects = calcQuantilesAndExpects()
    # print(expects)
    export_ans_to_excel(ans, expects)